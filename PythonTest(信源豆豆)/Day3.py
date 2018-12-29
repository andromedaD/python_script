import openpyxl

#打开excel文件
wb=openpyxl.load_workbook('new1.xlsx')
# 获取工作簿所有工作表名
print(wb.get_sheet_names())

#获取工作表
sheet = wb.get_sheet_by_name('aa')
print(sheet['A1'].value)

#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")


# 写入excel文件中 date 数据，date是list数据类型， fields 表头
def savetoexcel(data,fields,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入数据
        for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=wbname)
    print("保存成功")

savetoexcel('一二三四','2','aa','new1.xlsx');