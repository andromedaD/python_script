import openpyxl

# 写入excel文件中 date 数据，date是list数据类型， fields 表头
def savetoexcel(data,fields,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    row01=['/','用例名称','执行结果']

    for field in range(1,len(row01)+1):   # 写入表头
        _=sheet.cell(row=1, column=field, value=str(row01[field-1]))

    ss = 2;
    for field in range(1, len(row01) + 1):  # 写入表头
        _ = sheet.cell(row=ss, column=field, value=str(data[field - 1]))

    wb.save(filename=wbname)
    print("保存成功")

